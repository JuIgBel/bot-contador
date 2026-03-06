from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from database.db import agregar_gasto_db, obtener_gastos_db, obtener_todos_gastos_db, obtener_todas_facturas_db
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

ESPERANDO_GASTO = "esperando_gasto"
ESPERANDO_CATEGORIA = "esperando_categoria"
ESPERANDO_MONTO = "esperando_monto"

CATEGORIAS = ["🍔 Comida", "🏠 Hogar", "🚗 Transporte", "💊 Salud", "🎮 Entretenimiento", "👕 Ropa", "📚 Educación", "💼 Trabajo", "🔧 Servicios", "📦 Otros"]

async def agregar_gasto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    estado = context.user_data.get('estado')
    user_id = update.effective_user.id
    texto = update.message.text

    if estado == ESPERANDO_GASTO:
        context.user_data['gasto_desc'] = texto
        context.user_data['estado'] = ESPERANDO_CATEGORIA
        keyboard = [[InlineKeyboardButton(cat, callback_data=f"cat_{cat}")] for cat in CATEGORIAS]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "📂 *Seleccioná la categoría:*",
            parse_mode='Markdown',
            reply_markup=reply_markup
        )
    elif estado == ESPERANDO_MONTO:
        try:
            monto = float(texto.replace(',', '.').replace('$', '').strip())
            desc = context.user_data.get('gasto_desc')
            cat = context.user_data.get('gasto_cat')
            agregar_gasto_db(user_id, desc, cat, monto)
            context.user_data['estado'] = None
            await update.message.reply_text(
                f"✅ *Gasto registrado!*\n\n"
                f"📝 {desc}\n"
                f"📂 {cat}\n"
                f"💵 ${monto:,.2f}\n\n"
                f"Usá /menu para continuar.",
                parse_mode='Markdown'
            )
        except ValueError:
            await update.message.reply_text("❌ Ingresá un número válido (ej: `1500` o `1500.50`):", parse_mode='Markdown')

async def handle_categoria_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    categoria = query.data.replace("cat_", "")
    context.user_data['gasto_cat'] = categoria
    context.user_data['estado'] = ESPERANDO_MONTO
    await query.message.reply_text(
        f"📂 Categoría: *{categoria}*\n\n💵 Ingresá el monto (ej: `1500`):",
        parse_mode='Markdown'
    )

async def listar_gastos(update: Update, context: ContextTypes.DEFAULT_TYPE, via_callback=False):
    user_id = update.effective_user.id
    gastos = obtener_gastos_db(user_id)
    
    if not gastos:
        texto = "📋 No tenés gastos registrados todavía.\nUsá el menú para agregar uno."
    else:
        texto = "📋 *Últimos Gastos:*\n\n"
        for g in gastos:
            texto += f"• {g['fecha']} | {g['descripcion']} | {g['categoria']} | *${g['monto']:,.2f}*\n"
    
    if via_callback:
        await update.callback_query.message.reply_text(texto, parse_mode='Markdown')
    else:
        await update.message.reply_text(texto, parse_mode='Markdown')

async def exportar_excel(update: Update, context: ContextTypes.DEFAULT_TYPE, via_callback=False):
    user_id = update.effective_user.id
    gastos = obtener_todos_gastos_db(user_id)
    facturas = obtener_todas_facturas_db(user_id)
    
    wb = openpyxl.Workbook()
    
    # Hoja de Gastos
    ws_gastos = wb.active
    ws_gastos.title = "Gastos"
    headers_gastos = ["Fecha", "Descripción", "Categoría", "Monto"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, h in enumerate(headers_gastos, 1):
        cell = ws_gastos.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    total = 0
    for i, g in enumerate(gastos, 2):
        ws_gastos.cell(row=i, column=1, value=g['fecha'])
        ws_gastos.cell(row=i, column=2, value=g['descripcion'])
        ws_gastos.cell(row=i, column=3, value=g['categoria'])
        ws_gastos.cell(row=i, column=4, value=g['monto'])
        total += g['monto']
        if i % 2 == 0:
            for col in range(1, 5):
                ws_gastos.cell(row=i, column=col).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    
    fila_total = len(gastos) + 2
    ws_gastos.cell(row=fila_total, column=3, value="TOTAL").font = Font(bold=True)
    ws_gastos.cell(row=fila_total, column=4, value=total).font = Font(bold=True)
    
    for col in ws_gastos.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_gastos.column_dimensions[col[0].column_letter].width = max_length + 4
    
    # Hoja de Facturas
    ws_facturas = wb.create_sheet("Facturas")
    headers_facturas = ["Servicio", "Monto", "Vencimiento", "Estado"]
    for col, h in enumerate(headers_facturas, 1):
        cell = ws_facturas.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for i, f in enumerate(facturas, 2):
        ws_facturas.cell(row=i, column=1, value=f['servicio'])
        ws_facturas.cell(row=i, column=2, value=f['monto'])
        ws_facturas.cell(row=i, column=3, value=f['fecha_vencimiento'])
        ws_facturas.cell(row=i, column=4, value="✅ Pagada" if f['pagada'] else "⏳ Pendiente")
    
    for col in ws_facturas.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_facturas.column_dimensions[col[0].column_letter].width = max_length + 4
    
    filename = f"reporte_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = f"/tmp/{filename}"
    wb.save(filepath)
    
    if via_callback:
        chat = update.callback_query.message.chat
        await context.bot.send_document(
            chat_id=chat.id,
            document=open(filepath, 'rb'),
            filename=filename,
            caption="📊 *Tu reporte financiero*\n\nContiene gastos y facturas registradas.",
            parse_mode='Markdown'
        )
    else:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=open(filepath, 'rb'),
            filename=filename,
            caption="📊 *Tu reporte financiero*",
            parse_mode='Markdown'
        )
    os.remove(filepath)
